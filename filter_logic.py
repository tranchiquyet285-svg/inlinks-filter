"""
filter_logic.py
Core filtering logic — không phụ thuộc Streamlit.
"""

import io
import re
from collections import Counter
from typing import Optional, Union

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows


# ══════════════════════════════════════════════════════════════════
#  CẤU HÌNH — Universal (hoạt động trên mọi CMS/website)
# ══════════════════════════════════════════════════════════════════

# ── GIỮ: ít nhất 1 pattern phải khớp ──────────────────────────
CONTEXTUAL_PATTERNS: list = [

    # ── STRUCTURAL (universal — hoạt động trên mọi CMS) ──
    # Link nằm trong <p> = editorial text, đúng với mọi website
    r"/p(?:\[\d+\])?/a(?:\[\d+\])?$",
    r"/p(?:\[\d+\])?/(?:strong|em|b|i|u|span|cite)(?:\[\d+\])?/a(?:\[\d+\])?$",
    r"/p(?:\[\d+\])?/a(?:\[\d+\])?/(?:strong|em|b|span)(?:\[\d+\])?$",
    # Link trực tiếp trong <li> (list trong body text)
    r"/li(?:\[\d+\])?/a(?:\[\d+\])?$",
    r"/li(?:\[\d+\])?/(?:strong|em|b|i|u|span)(?:\[\d+\])?/a(?:\[\d+\])?$",
    # Link trong <li> chứa <p>
    r"/li(?:\[\d+\])?/p(?:\[\d+\])?/a(?:\[\d+\])?$",

    # ── WORDPRESS / CLASSIC THEMES ──
    r"[@/]id=['\"]post-\d+['\"]",
    r"@class=['\"][^'\"]*\bentry-content\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bpost-content\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bsingle-content\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\barticle-content\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bthe-content\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bcontent-area\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bpage-content\b[^'\"]*['\"]",
    r"/article(?:\[@[^\]]*\])?/",

    # ── WOOCOMMERCE ──
    r"[@/]id=['\"]tab-description['\"]",
    r"@class=['\"][^'\"]*\bwoocommerce-Tabs-panel--description\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bwoocommerce-product-details__short-description\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bproduct-description\b[^'\"]*['\"]",

    # ── CATEGORY / ARCHIVE DESCRIPTION ──
    r"@class=['\"][^'\"]*\bterm-description\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bcategory-description\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\barchive-description\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\btaxonomy-description\b[^'\"]*['\"]",

    # ── POPULAR THEMES & PAGE BUILDERS ──
    r"@class=['\"][^'\"]*\belementor-widget-text-editor\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bwp-block-paragraph\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bwp-block-list\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\brich-text\b[^'\"]*['\"]",

    # ── XENANGPLUS.COM SPECIFIC (giữ lại để tương thích) ──
    r"/main/div\[\d+\]/div/(?:p(?:\[\d+\])?|ul(?:\[\d+\])?|strong|span|em|b|h[1-6])(?:/|\[|$)",
]

# LƯU Ý: Pattern /article/ đã bị xoá vì quá rộng — bắt nhầm sidebar/widget
# nằm trong <article>. Link editorial thật sự luôn đi qua <p> tag,
# đã được bắt bởi pattern /p/a$ phía trên.

# ── LOẠI: không được khớp bất kỳ pattern nào ──────────────────
EXCLUDE_PATTERNS: dict = {
    # Navigation & UI
    r"[@/]id=['\"]tab-title-":                   "tab-title (chuyển tab)",
    r"[@/]id=['\"]comments['\"]":                "khu vực comment",
    r"[@/]id=['\"]respond['\"]":                 "form comment",
    r"[@/]id=['\"]login['\"]":                   "trang đăng nhập",
    r"[@/]id=['\"]backtoblog['\"]":              "link back-to-blog",
    r"[@/]id=['\"]ez-toc-container['\"]":        "mục lục TOC",
    r"@class=['\"][^'\"]*\btoc\b[^'\"]*['\"]":  "mục lục TOC (class)",
    # Product grids & cards
    r"[@/]id=['\"]products['\"]":                "grid sản phẩm (id=products)",
    r"@class=['\"]block['\"]":                   "card sản phẩm (class=block)",
    r"@class=['\"][^'\"]*\bproducts\b[^'\"]*['\"]": "grid sản phẩm (class)",
    r"@class=['\"][^'\"]*\bwoocommerce-loop-product\b[^'\"]*['\"]": "card loop WooCommerce",
    r"@class=['\"][^'\"]*\brelated\b[^'\"]*['\"]":  "sản phẩm liên quan",
    # Sliders & carousels
    r"swiper-slide":                              "carousel/slider",
    r"@class=['\"][^'\"]*\bslider\b[^'\"]*['\"]": "slider",
    r"@class=['\"][^'\"]*\bcarousel\b[^'\"]*['\"]": "carousel",
    # Tables (product listings)
    r"/table/tbody/":                             "bảng listing sản phẩm",
    # Breadcrumb & pagination
    r"@class=['\"][^'\"]*\bbreadcrumb\b[^'\"]*['\"]": "breadcrumb",
    r"@class=['\"][^'\"]*\bpagination\b[^'\"]*['\"]": "pagination",
    r"@class=['\"][^'\"]*\bpage-numbers\b[^'\"]*['\"]": "page numbers",
    # Subcategory / topic navigation boxes (SEOSONA & similar)
    r"@class=['\"][^'\"]*\bbox-subcategoryseo\b[^'\"]*['\"]": "box subcategory nav",
    r"@class=['\"][^'\"]*\bbox-subcategor\b[^'\"]*['\"]":     "box subcategory nav",
    # Divi / page builder layout blocks (id="section_123", id="col-456")
    r"[@/]id=['\"]section_\d+['\"]":  "Divi section block",
    r"[@/]id=['\"]col-\d+['\"]":      "Divi column block",
    r"[@/]id=['\"]row-\d+['\"]":      "Divi row block",
    r"[@/]id=['\"]module_\d+['\"]":   "Divi module block",
}

BUTTON_ANCHORS: frozenset = frozenset({
    "xem thông tin", "xem tất cả", "xem ngay", "xem chi tiết",
    "toggle", "trang sau", "trang trước", "«", "»", "‹", "›",
    "next", "prev", "previous", "load more", "read more",
    "mua ngay", "thêm vào giỏ",
})

GROUP_COLORS: dict = {
    "Tab Mô tả sản phẩm": "DBEAFE",
    "Body bài viết":       "D1FAE5",
    "Mô tả category":      "FEF3C7",
    "Paragraph link":      "E0F2FE",
    "Page content":        "FCE7F3",
    "Contextual (khác)":   "F3F4F6",
}

HEADER_BG = "1D4ED8"


# ══════════════════════════════════════════════════════════════════
#  CHUẨN HOÁ DATAFRAME
# ══════════════════════════════════════════════════════════════════

COLUMN_ALIASES: dict = {
    "Anchor":      "Anchor Text",
    "Alt Text":    "Anchor Text",
    "Position":    "Link Position",
    "Path":        "Link Path",
    "Source":      "From",
    "Destination": "To",
    "Target":      "To",
}

REQUIRED_COLS = {"Type", "Link Position", "Anchor Text", "Link Path", "From", "To"}


def normalise(df: pd.DataFrame) -> tuple:
    """
    Chuẩn hoá tên cột.
    Trả về (df, error_message) — error_message là None nếu OK.
    """
    df = df.copy()
    df.columns = [str(c).strip() for c in df.columns]
    for old, new in COLUMN_ALIASES.items():
        if old in df.columns and new not in df.columns:
            df = df.rename(columns={old: new})
    missing = REQUIRED_COLS - set(df.columns)
    if missing:
        return None, f"File thiếu cột: {missing}. Cột có trong file: {list(df.columns)}"
    for col in REQUIRED_COLS:
        df[col] = df[col].fillna("").astype(str).str.strip()
    return df, None


# ══════════════════════════════════════════════════════════════════
#  LOGIC LỌC
# ══════════════════════════════════════════════════════════════════

# Container nội dung "mạnh" — nếu link nằm trong đây thì bỏ qua
# exclude patterns (ví dụ: bảng giá trong mô tả category vẫn được giữ)
STRONG_CONTENT_PATTERNS: list = [
    r"[@/]id=['\"]tab-description['\"]",
    r"[@/]id=['\"]post-\d+['\"]",
    r"@class=['\"][^'\"]*\bentry-content\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bpost-content\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bsingle-content\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bthe-content\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bterm-description\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bcategory-description\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\barchive-description\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\btaxonomy-description\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bwoocommerce-Tabs-panel--description\b[^'\"]*['\"]",
    r"@class=['\"][^'\"]*\bwoocommerce-product-details__short-description\b[^'\"]*['\"]",
]


def _matches_any(patterns: Union[list, dict], text: str) -> bool:
    keys = patterns.keys() if isinstance(patterns, dict) else patterns
    return any(re.search(p, text, re.IGNORECASE) for p in keys)


def _is_button(anchor: str) -> bool:
    a = anchor.strip()
    return a.lower() in BUTTON_ANCHORS or bool(re.fullmatch(r"\d+", a))


def classify_group(path: str) -> str:
    # WooCommerce product description tab
    if re.search(r"[@/]id=['\"]tab-description['\"]", path, re.I):
        return "Tab Mô tả sản phẩm"
    if re.search(r"woocommerce-Tabs-panel--description|product-description|woocommerce-product-details", path, re.I):
        return "Tab Mô tả sản phẩm"
    # WordPress post / page body
    if re.search(r"[@/]id=['\"]post-\d+['\"]", path, re.I):
        return "Body bài viết"
    if re.search(r"entry-content|post-content|single-content|article-content|the-content|page-content", path, re.I):
        return "Body bài viết"
    # Category / archive description
    if re.search(r"term-description|category-description|archive-description|taxonomy-description", path, re.I):
        return "Mô tả category"
    # Structural: link trong paragraph
    if re.search(r"/p(?:\[\d+\])?/(?:(?:strong|em|b|i|u|span|cite)(?:\[\d+\])?/)?a(?:\[\d+\])?$", path, re.I):
        return "Paragraph link"
    # Page builders
    if re.search(r"elementor|wp-block|rich-text", path, re.I):
        return "Body bài viết"
    return "Contextual (khác)"


def apply_filter(df: pd.DataFrame) -> tuple:
    m_type       = df["Type"] == "Hyperlink"
    m_pos        = df["Link Position"] == "Content"
    m_anchor     = df["Anchor Text"] != ""
    m_no_btn     = ~df["Anchor Text"].apply(_is_button)
    # Link trong strong content container → bỏ qua exclude (bảng giá trong mô tả, v.v.)
    m_strong     = df["Link Path"].apply(lambda p: _matches_any(STRONG_CONTENT_PATTERNS, p))
    m_no_excl    = ~df["Link Path"].apply(lambda p: _matches_any(EXCLUDE_PATTERNS, p))
    m_pass_excl  = m_no_excl | m_strong   # strong container overrides exclude
    m_cx         = df["Link Path"].apply(lambda p: _matches_any(CONTEXTUAL_PATTERNS, p))

    kept_mask = m_type & m_pos & m_anchor & m_no_btn & m_pass_excl & m_cx

    stats = {
        "total":         len(df),
        "non_hyperlink": int((~m_type).sum()),
        "non_content":   int((m_type & ~m_pos).sum()),
        "empty_anchor":  int((m_type & m_pos & ~m_anchor).sum()),
        "button_anchor": int((m_type & m_pos & m_anchor & ~m_no_btn).sum()),
        "excluded_path": int((m_type & m_pos & m_anchor & m_no_btn & ~m_pass_excl).sum()),
        "no_match":      int((m_type & m_pos & m_anchor & m_no_btn & m_pass_excl & ~m_cx).sum()),
        "kept":          int(kept_mask.sum()),
        "groups":        Counter(),
    }

    kept = df[kept_mask].copy()
    kept["Group"] = kept["Link Path"].apply(classify_group)
    stats["groups"] = Counter(kept["Group"])
    return kept, stats


def debug_page(df: pd.DataFrame, url: str) -> pd.DataFrame:
    """
    Trả về tất cả link từ một URL (From chứa url),
    kèm cột Kết_Quả cho biết GIỮ hay lý do bị loại.
    Dùng để debug tại sao link bị lọc mất.
    """
    sub = df[df["From"].str.contains(url, case=False, na=False)].copy()
    results = []
    for _, row in sub.iterrows():
        t      = row["Type"]
        pos    = row["Link Position"]
        anchor = row["Anchor Text"]
        path   = row["Link Path"]

        if t != "Hyperlink":
            results.append(f"❌ Type={t!r}")
        elif pos != "Content":
            results.append(f"❌ Position={pos!r}")
        elif not anchor:
            results.append("❌ Anchor rỗng")
        elif _is_button(anchor):
            results.append(f"❌ Button: {anchor!r}")
        else:
            is_strong = _matches_any(STRONG_CONTENT_PATTERNS, path)
            excl_reason = next(
                (label for pat, label in EXCLUDE_PATTERNS.items()
                 if re.search(pat, path, re.IGNORECASE)),
                None,
            )
            if excl_reason and not is_strong:
                results.append(f"❌ Loại: {excl_reason}")
            elif excl_reason and is_strong:
                results.append(f"✅ GIỮ — strong container override ({excl_reason} bỏ qua)")
            elif not _matches_any(CONTEXTUAL_PATTERNS, path):
                results.append("❌ Không khớp contextual")
            else:
                results.append(f"✅ GIỮ ({classify_group(path)})")

    sub = sub.copy()
    sub["Kết_Quả"] = results
    cols = ["Kết_Quả", "To", "Anchor Text", "Link Path", "Link Position"]
    return sub[[c for c in cols if c in sub.columns]]


# ══════════════════════════════════════════════════════════════════
#  EXCEL BUILDER
# ══════════════════════════════════════════════════════════════════

def _style_header(ws):
    fill  = PatternFill("solid", fgColor=HEADER_BG)
    font  = Font(bold=True, color="FFFFFF", name="Calibri", size=11)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    for cell in ws[1]:
        cell.fill = fill
        cell.font = font
        cell.alignment = align
    ws.row_dimensions[1].height = 30


def _autofit(ws, max_w: int = 70):
    for col in ws.columns:
        w = max(len(str(c.value or "")) for c in col)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(w + 4, max_w)


def _write_df(ws, df: pd.DataFrame):
    for row in dataframe_to_rows(df, index=False, header=True):
        ws.append(row)
    _style_header(ws)
    if df.shape[0] > 0:
        ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"
    _autofit(ws)


def _color_by_group(ws):
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        gval  = str(row[0].value or "")
        color = GROUP_COLORS.get(gval, "FFFFFF")
        fill  = PatternFill("solid", fgColor=color)
        for cell in row:
            cell.fill = fill


def build_excel(kept: pd.DataFrame, stats: dict) -> bytes:
    """Xuất Excel 4 sheet, trả về bytes để dùng với st.download_button."""
    wb = Workbook()
    wb.remove(wb.active)

    # Sheet 1: Links đã lọc
    ws1 = wb.create_sheet("Contextual Links")
    priority = ["Group", "From", "To", "Anchor Text", "Link Path",
                "Status Code", "Follow", "Type", "Link Position"]
    cols = [c for c in priority if c in kept.columns] + \
           [c for c in kept.columns if c not in priority]
    _write_df(ws1, kept[cols])
    _color_by_group(ws1)

    # Sheet 2: Theo From
    ws2 = wb.create_sheet("Theo From (nguồn)")
    if len(kept) > 0:
        from_df = (
            kept.groupby("From", as_index=False)
            .agg(So_Link_Ra=("To", "count"), So_Trang_Dich_Rieng=("To", "nunique"))
            .sort_values("So_Link_Ra", ascending=False)
        )
        from_df.columns = ["From", "Số Link Ra", "Số Trang Đích Riêng"]
    else:
        from_df = pd.DataFrame(columns=["From", "Số Link Ra", "Số Trang Đích Riêng"])
    _write_df(ws2, from_df)

    # Sheet 3: Theo To
    ws3 = wb.create_sheet("Theo To (đích — SEO)")
    if len(kept) > 0:
        to_df = (
            kept.groupby("To", as_index=False)
            .agg(So_Link_Vao=("From", "count"), So_Trang_Nguon_Rieng=("From", "nunique"))
            .sort_values("So_Link_Vao", ascending=False)
        )
        to_df.columns = ["To", "Số Link Vào", "Số Trang Nguồn Riêng"]
    else:
        to_df = pd.DataFrame(columns=["To", "Số Link Vào", "Số Trang Nguồn Riêng"])
    _write_df(ws3, to_df)

    # Sheet 4: Quy tắc & Thống kê
    ws4 = wb.create_sheet("Quy tắc & Thống kê")
    pct = f"{stats['kept'] / stats['total'] * 100:.1f}%" if stats["total"] else "0%"
    rows_data = [
        ("CHỈ SỐ", "GIÁ TRỊ"),
        ("Tổng link input", stats["total"]),
        ("Giữ lại (contextual)", stats["kept"]),
        ("Tỷ lệ giữ lại", pct),
        ("", ""),
        ("LÝ DO LOẠI", "SỐ LƯỢNG"),
        ("Type ≠ Hyperlink",             stats["non_hyperlink"]),
        ("Position ≠ Content",           stats["non_content"]),
        ("Anchor rỗng",                  stats["empty_anchor"]),
        ("Anchor là button / số",        stats["button_anchor"]),
        ("Path chứa marker loại trừ",    stats["excluded_path"]),
        ("Không khớp contextual",        stats["no_match"]),
        ("", ""),
        ("NHÓM GIỮ LẠI", "SỐ LƯỢNG"),
    ]
    for grp, cnt in sorted(stats["groups"].items(), key=lambda x: -x[1]):
        rows_data.append((grp, cnt))

    rows_data += [("", ""), ("CONTEXTUAL PATTERNS (GIỮ)", "")]
    for p in CONTEXTUAL_PATTERNS:
        rows_data.append((p, ""))
    rows_data += [("", ""), ("EXCLUDE PATTERNS (LOẠI)", "")]
    for p, label in EXCLUDE_PATTERNS.items():
        rows_data.append((p, label))

    section_headers = {
        "CHỈ SỐ", "LÝ DO LOẠI", "NHÓM GIỮ LẠI",
        "CONTEXTUAL PATTERNS (GIỮ)", "EXCLUDE PATTERNS (LOẠI)",
    }
    for i, (a, b) in enumerate(rows_data, start=1):
        ws4.cell(i, 1, a)
        ws4.cell(i, 2, b)
        if a in section_headers:
            for col_idx in (1, 2):
                c = ws4.cell(i, col_idx)
                c.fill = PatternFill("solid", fgColor=HEADER_BG)
                c.font = Font(bold=True, color="FFFFFF", name="Calibri")
    ws4.column_dimensions["A"].width = 55
    ws4.column_dimensions["B"].width = 22

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
